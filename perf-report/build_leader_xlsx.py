#!/usr/bin/env python3
"""Build per_leader_breakdown.xlsx from leader_slots_24h.csv with src_ip and
ASN/datacenter columns (looked up via ipinfo.io, cached in DATA_DIR).

Inputs in DATA_DIR:
  leader_slots_24h.csv  from extract.py
  meta.json             from extract.py
  name_ip_map.txt       "name|src_ip" lines, dump of shred_providers:
                        psql "$ASTRALANE_DB_URL" -Atc \
                          "SELECT name, src_ip FROM shred_providers" \
                          > $DATA_DIR/name_ip_map.txt
  ip_asn.json           lookup cache (created/updated automatically)

Environment:
  DATA_DIR  optional  default ./data
  OUT_DIR   optional  default ./out
"""

import csv
import json
import os
import time
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = os.environ.get("DATA_DIR", "./data")
OUT_DIR = os.environ.get("OUT_DIR", "./out")
OUT = f"{OUT_DIR}/per_leader_breakdown.xlsx"
HUB = {"198.13.138.175": "fra (198.13.138.175)",
       "64.130.37.201": "ny (64.130.37.201)"}

name_ip = {}
for line in open(f"{SRC}/name_ip_map.txt"):
    line = line.strip()
    if line:
        name, ip = line.split("|")
        name_ip.setdefault(name, ip)


def ip_of(source):
    ip = name_ip.get(source)
    if not ip and (source.startswith("fra-") or source.startswith("ny-")):
        cand = source.split("-", 1)[1]
        if cand.count(".") == 3:
            ip = cand
    return ip or ""


try:
    asn = json.load(open(f"{SRC}/ip_asn.json"))
except FileNotFoundError:
    asn = {}

rows = list(csv.DictReader(open(f"{SRC}/leader_slots_24h.csv")))
meta = json.load(open(f"{SRC}/meta.json"))
window = f"{meta['window'][0][:16]} -> {meta['window'][1][:16]} UTC"

# fetch ASN for any IPs not cached yet
for r in rows:
    ip = ip_of(r["source"])
    if ip and ip not in asn:
        try:
            with urllib.request.urlopen(f"https://ipinfo.io/{ip}/json",
                                        timeout=15) as resp:
                d = json.load(resp)
            asn[ip] = {"org": d.get("org", ""), "city": d.get("city", ""),
                       "country": d.get("country", "")}
            print("ipinfo", ip, "->", asn[ip]["org"])
            time.sleep(0.4)
        except Exception as e:
            asn[ip] = {"org": f"lookup failed: {e}", "city": "", "country": ""}
json.dump(asn, open(f"{SRC}/ip_asn.json", "w"), indent=1)

os.makedirs(OUT_DIR, exist_ok=True)
wb = Workbook()
ws = wb.active
ws.title = "per_leader_24h"
ws.cell(1, 1, f"Per-leader breakdown — last 24h (window {window})") \
    .font = Font(bold=True, size=13)
ws.cell(2, 1, "For FEC sets in slots led by each provider-validator: how "
              "every source performed. delta vs jito baseline; negative = "
              "earlier. ASN/DC from ipinfo.io.")
hdr = ["hub", "leader_provider", "slots_seen", "fec_sets", "source", "src_ip",
       "asn_dc", "dc_location", "samples", "wins", "win_rate_pct",
       "avg_delta_ms"]
for c, h in enumerate(hdr, 1):
    cell = ws.cell(4, c, h)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
ws.freeze_panes = "A5"

rows.sort(key=lambda r: (r["hub"], r["leader_provider"],
                         -float(r["win_rate_pct"])))
r_i = 5
for r in rows:
    ip = ip_of(r["source"])
    a = asn.get(ip, {})
    loc = f"{a.get('city', '')}, {a.get('country', '')}".strip(", ")
    vals = [HUB.get(r["hub"], r["hub"]), r["leader_provider"],
            int(r["slots_seen"]), int(r["fec_sets"]), r["source"], ip,
            a.get("org", ""), loc, int(r["samples"]), int(r["wins"]),
            float(r["win_rate_pct"]), float(r["avg_delta_ms"])]
    for c, v in enumerate(vals, 1):
        ws.cell(r_i, c, v)
    r_i += 1

for col in ws.columns:
    width = max((len(str(c.value)) for c in col if c.value is not None),
                default=8)
    ws.column_dimensions[get_column_letter(col[0].column)].width = \
        min(width + 2, 44)
ws.auto_filter.ref = f"A4:L{r_i - 1}"
wb.save(OUT)
print("saved", OUT, "rows:", r_i - 5)
