#!/usr/bin/env python3
"""Build the general shred-source performance xlsx plus one small shareable
xlsx per provider, from the CSVs produced by extract.py.

Environment:
  DATA_DIR  optional  dir with extract.py CSVs (default ./data)
  OUT_DIR   optional  output dir (default ./out)

Outputs:
  $OUT_DIR/shred_source_performance.xlsx
  $OUT_DIR/provider_reports/shred_perf_<provider>.xlsx
"""

import csv
import json
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = os.environ.get("DATA_DIR", "./data")
OUT_DIR = os.environ.get("OUT_DIR", "./out")
OUT = f"{OUT_DIR}/shred_source_performance.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GOOD = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="FFC7CE")

HUB_LABELS = {
    "198.13.138.175": "fra (198.13.138.175)",
    "64.130.37.201": "ny (64.130.37.201)",
}


def hub_label(v):
    return HUB_LABELS.get(str(v), v)


def brand_of(source):
    s = source.lower()
    for prefix, b in (("everstake", "everstake"), ("lucky", "lucky-stake"),
                      ("twinstake", "twinstake"), ("jupiter", "jupiter"),
                      ("thor", "thor"), ("dawn", "dawn-labs"),
                      ("assymetric", "asymmetric"), ("asymmetric", "asymmetric"),
                      ("staking-fac", "staking-facilities"), ("p2p", "p2p.org"),
                      ("jito", "jito"), ("soyas", "soyas"),
                      ("soldiver", "soldiver")):
        if s.startswith(prefix):
            return b
    return "other"


def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f == int(f) and "." not in str(v) else f
    except ValueError:
        return v


def add_csv_sheet(wb, title, path, autofilter=True):
    ws = wb.create_sheet(title)
    with open(path) as f:
        rdr = csv.reader(f)
        header = next(rdr)
        hub_col = header.index("hub") if "hub" in header else None
        for c, val in enumerate(header, 1):
            ws.cell(1, c, val)
        for r, row in enumerate(rdr, 2):
            for c, val in enumerate(row, 1):
                if hub_col is not None and c == hub_col + 1:
                    ws.cell(r, c, hub_label(val))
                else:
                    ws.cell(r, c, num(val))
    style_header(ws)
    if autofilter and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    autosize(ws)
    return ws


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def autosize(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None),
                    default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(width + 2, 42)


def h2h_matrices(wb, path="head_to_head_24h.csv"):
    rows = list(csv.DictReader(open(f"{SRC}/{path}")))
    hubs = sorted({r["hub"] for r in rows})
    for hub in hubs:
        pct = {}
        srcs = set()
        for r in rows:
            if r["hub"] != hub:
                continue
            a, b = r["source_a"], r["source_b"]
            srcs.update((a, b))
            n = int(r["both_present"])
            af = int(r["a_first_count"])
            pct[(a, b)] = 100 * af / n if n else None
            pct[(b, a)] = 100 * (n - af) / n if n else None

        def score(s):
            vals = [pct[(s, o)] for o in srcs if o != s and (s, o) in pct]
            return -(sum(vals) / len(vals)) if vals else 0

        order = sorted(srcs, key=score)
        ws = wb.create_sheet(f"h2h {hub_label(hub)}"[:31])
        ws.cell(1, 1, "row beats col %")
        for j, b in enumerate(order, 2):
            ws.cell(1, j, b)
        for i, a in enumerate(order, 2):
            ws.cell(i, 1, a).font = Font(bold=True)
            for j, b in enumerate(order, 2):
                if a == b:
                    continue
                v = pct.get((a, b))
                if v is None:
                    continue
                cell = ws.cell(i, j, round(v, 1))
                if v >= 60:
                    cell.fill = GOOD
                elif v <= 40:
                    cell.fill = BAD
        style_header(ws)
        ws.freeze_panes = "B2"
        autosize(ws)


def trend_charts(wb, trend_ws):
    """Pivot daily_trend per hub and draw line charts onto the trend sheet."""
    rows = list(csv.DictReader(open(f"{SRC}/daily_trend_7d.csv")))
    hubs = sorted({r["hub"] for r in rows})
    pivot_ws = wb.create_sheet("trend_pivots")
    pivot_ws.sheet_state = "hidden"

    anchor_row = 2
    pcol = 1
    for hub in hubs:
        hrows = [r for r in rows if r["hub"] == hub]
        days = sorted({r["day"] for r in hrows})
        sources = sorted({r["source"] for r in hrows})
        for metric, title in (
                ("avg_delta_ms", "avg delta vs jito, ms (neg = earlier)"),
                ("beats_jito_pct", "% of samples beating jito")):
            vals = {(r["day"], r["source"]): r[metric] for r in hrows}
            top = pivot_ws.cell(1, pcol, f"{hub_label(hub)} {metric}").row
            for j, s in enumerate(sources):
                pivot_ws.cell(top + 1, pcol + 1 + j, s)
            for i, d in enumerate(days):
                pivot_ws.cell(top + 2 + i, pcol, d)
                for j, s in enumerate(sources):
                    v = vals.get((d, s))
                    if v not in (None, ""):
                        pivot_ws.cell(top + 2 + i, pcol + 1 + j, float(v))
            chart = LineChart()
            chart.title = f"{hub_label(hub)} — {title}"
            chart.height = 10
            chart.width = 26
            chart.y_axis.title = metric
            chart.x_axis.title = "day"
            data = Reference(pivot_ws, min_col=pcol + 1,
                             max_col=pcol + len(sources),
                             min_row=top + 1, max_row=top + 1 + len(days))
            cats = Reference(pivot_ws, min_col=pcol, min_row=top + 2,
                             max_row=top + 1 + len(days))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            trend_ws.add_chart(chart, f"H{anchor_row}")
            anchor_row += 21
            pcol += len(sources) + 2


# Columns shown on shareable panels (avg_rank omitted — it hints at how many
# competing streams we ingest).
PANEL_OVERALL_COLS = ["hub", "source", "samples", "win_rate_pct",
                      "avg_delta_ms", "p50_ms", "p90_ms", "p99_ms",
                      "beats_jito_pct", "fec_set_coverage_pct"]
PANEL_OWN_COLS = ["hub", "source", "slots_kind", "samples",
                  "win_rate_of_samples_pct", "avg_delta_ms"]


def panel_sheet(wb, b, meta, per_source, own):
    """Sanitized sheet for one provider: only their own streams' stats —
    shareable with that provider without exposing other sources/IPs."""
    ws = wb.create_sheet(f"panel {b}"[:31])
    ws.cell(1, 1, f"{b} — shred delivery performance, "
                  f"last {meta['window_hours']}h").font = Font(bold=True, size=13)
    ws.cell(2, 1, f"window: {meta['window'][0]} .. {meta['window'][1]} "
                  f"| hubs: fra, ny")
    ws.cell(3, 1, "delta = arrival vs the jito baseline on the same hub for "
                  "the same FEC set; NEGATIVE = earlier than jito.")
    ws.cell(4, 1, "win_rate = % of sampled FEC sets where your stream "
                  "delivered first. coverage = % of sampled FEC sets your "
                  "stream delivered at all.")
    r = 6
    ws.cell(r, 1, "OVERALL").font = Font(bold=True)
    r += 1
    for c, col in enumerate(PANEL_OVERALL_COLS, 1):
        cell = ws.cell(r, c, col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    for row in per_source:
        if brand_of(row["source"]) != b:
            continue
        r += 1
        for c, col in enumerate(PANEL_OVERALL_COLS, 1):
            v = hub_label(row[col]) if col == "hub" else num(row[col])
            ws.cell(r, c, v)
    r += 2
    own_rows = [x for x in own if brand_of(x["source"]) == b]
    if own_rows:
        ws.cell(r, 1, "YOUR LEADER SLOTS vs OTHER SLOTS").font = Font(bold=True)
        r += 1
        ws.cell(r, 1, "(own = slots where your validator was leader; "
                      "win_rate here is per your delivered samples)")
        r += 1
        for c, col in enumerate(PANEL_OWN_COLS, 1):
            cell = ws.cell(r, c, col)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
        for row in own_rows:
            r += 1
            for c, col in enumerate(PANEL_OWN_COLS, 1):
                v = hub_label(row[col]) if col == "hub" else num(row[col])
                ws.cell(r, c, v)
    for col_letter, width in (("A", 26), ("B", 24), ("C", 12), ("D", 12),
                              ("E", 14), ("F", 10), ("G", 10), ("H", 10),
                              ("I", 14), ("J", 18)):
        ws.column_dimensions[col_letter].width = width


def build_provider_files(meta):
    """One small xlsx per provider, safe to forward to them as-is."""
    outdir = f"{OUT_DIR}/provider_reports"
    os.makedirs(outdir, exist_ok=True)
    for f in os.listdir(outdir):
        os.remove(f"{outdir}/{f}")
    per_source = list(csv.DictReader(open(f"{SRC}/per_source_24h.csv")))
    own = list(csv.DictReader(open(f"{SRC}/own_vs_other_24h.csv")))
    brands = sorted({brand_of(r["source"]) for r in per_source}
                    - {"other", "jito"})
    files = []
    for b in brands:
        wb = Workbook()
        wb.remove(wb.active)
        panel_sheet(wb, b, meta, per_source, own)
        path = f"{outdir}/shred_perf_{b.replace('.', '_')}.xlsx"
        wb.save(path)
        files.append(path)
    return files


def readme(wb, meta):
    ws = wb.create_sheet("README", 0)
    lines = [
        ("Shred source performance report", ""),
        ("generated (UTC)", meta["generated_utc"]),
        ("main window", f"last {meta['window_hours']}h: "
                        f"{meta['window'][0]} .. {meta['window'][1]}"),
        ("epochs (leader schedule via Solana RPC)",
         ", ".join(map(str, meta["epochs"]))),
        ("FEC sets analyzed per hub", json.dumps(meta["fec_sets_per_hub"])),
        ("", ""),
        ("DEFINITIONS", ""),
        ("delta_ms", "arrival time vs the jito baseline source on the same "
                     "hub for the same FEC set; NEGATIVE = earlier than jito"),
        ("win_rate", "% of analyzed FEC sets where this source was the FIRST "
                     "of all sources that delivered it"),
        ("avg_rank", "average arrival position within a FEC set (0 = first)"),
        ("beats_jito_pct", "% of samples with delta < 0"),
        ("fec_set_coverage_pct", "% of analyzed FEC sets in which the source "
                                 "appeared at all (low = misses sets, also "
                                 "caps win_rate)"),
        ("h2h sheets", "row beats col: % of FEC sets both delivered where "
                       "the row source arrived first"),
        ("p50/p90/p99", "percentiles of delta_ms (20us histogram resolution)"),
        ("", ""),
        ("HUBS", ""),
        ("198.13.138.175", "terra-1 (FRA prod hub)"),
        ("64.130.37.201", "NY hub"),
        ("merry-gar", "test hub on merry-gar (88.216.36.99): astra vs lucky "
                      "vs jito vs everstake comparison; 'astra' = our own "
                      "source; delta baseline = jito on that hub"),
        ("", ""),
        ("CAVEATS", ""),
        ("sampling", "competition records 1-in-N FEC sets; stats are over "
                     "sampled sets, not every shred"),
        ("synthesized rows", "rows with delta_ns NULL (expected source never "
                             "arrived) are EXCLUDED from latency stats but "
                             "reflected in fec_set_coverage"),
        ("leader mapping", f"slots with unknown leader in window: "
                           f"{meta['unknown_leader_slots']}"),
        ("", ""),
        ("LEADER-PROVIDER IDENTITY PUBKEYS USED", ""),
    ]
    for prov, keys in meta["provider_pubkeys"].items():
        lines.append((prov, ", ".join(keys)))
    for r, (a, b) in enumerate(lines, 1):
        ws.cell(r, 1, a).font = Font(bold=bool(b == "" or a.isupper() or r == 1))
        ws.cell(r, 2, b)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 120


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    meta = json.load(open(f"{SRC}/meta.json"))
    wb = Workbook()
    wb.remove(wb.active)
    readme(wb, meta)
    add_csv_sheet(wb, "per_source_24h", f"{SRC}/per_source_24h.csv")
    h2h_matrices(wb)
    if os.path.exists(f"{SRC}/merry_per_source_24h.csv"):
        add_csv_sheet(wb, "astra vs lucky vs jito (merry)",
                      f"{SRC}/merry_per_source_24h.csv")
        h2h_matrices(wb, "merry_h2h_24h.csv")
    add_csv_sheet(wb, "leader_slots_24h", f"{SRC}/leader_slots_24h.csv")
    add_csv_sheet(wb, "own_vs_other_slots", f"{SRC}/own_vs_other_24h.csv")
    trend_ws = add_csv_sheet(wb, "daily_trend_7d", f"{SRC}/daily_trend_7d.csv")
    trend_charts(wb, trend_ws)
    add_csv_sheet(wb, "first_provider", f"{SRC}/first_provider.csv")
    add_csv_sheet(wb, "ip_coverage", f"{SRC}/ip_coverage.csv")
    wb.save(OUT)
    print("saved", OUT)
    for p in build_provider_files(meta):
        print("saved", p)


if __name__ == "__main__":
    main()
